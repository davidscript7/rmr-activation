#!/usr/bin/env python3

"""
============================================================
RMR ACTIVATION — STEP 3: CONTRACT DETAIL → RMR ACTIVATION.xlsx
Reads the SQ01 report, processes dashes in Legacy Contract
Number, merges with IW75 data to obtain CFIN DATE, and
APPENDS the final data to the shared RMR ACTIVATION.xlsx file,
distributing records equitably between Hector and Daniel.
============================================================
"""

import os
import sys
import json
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────
# CONFIGURATION — Update column names here if the
# SQ01 exported Excel changes its structure
# ─────────────────────────────────────────────────────────

SCRIPT_DIR    = os.path.dirname(os.path.abspath(__file__))
INPUT_FOLDER  = os.path.join(SCRIPT_DIR, "input")
OUTPUT_FOLDER = os.path.join(SCRIPT_DIR, "output")
INTER_STEP2   = os.path.join(INPUT_FOLDER, ".step2_data.json")

CONFIG_PATH = os.path.join(SCRIPT_DIR, "config.json")

# CONTRACT DETAIL column names (from SQ01 export)
COL_LEGACY_NUM  = "Legacy contract number"
COL_LEGACY_LINE = "Legacy contract Line number"
COL_TRANS_ID    = "Transaction ID"
COL_ITEM_NUM    = "Item Number in Document"

# ── Owners ────────────────────────────────────────────────
OWNER_1 = "Hector"
OWNER_2 = "Daniel"

# Local path for the shared RMR ACTIVATION file
LOCAL_RMR = os.path.join(OUTPUT_FOLDER, "RMR ACTIVATION.xlsx")

def banner(text):
    print("\n" + "=" * 60)
    print(f"  {text}")
    print("=" * 60)


# ─────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────

def load_config():
    if not os.path.exists(CONFIG_PATH):
        print("  ❌ config.json not found.")
        print(f"     Expected at: {CONFIG_PATH}")
        print("     Copy config.example.json → config.json and fill in your values.")
        sys.exit(1)
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


# ─────────────────────────────────────────────────────────
# CONTRACT DETAIL — file reading
# ─────────────────────────────────────────────────────────

def find_contract_file():
    files = sorted(
        [f for f in os.listdir(INPUT_FOLDER)
         if f.upper().startswith("CONTRACT DETAIL") and f.endswith(".xlsx")],
        reverse=True
    )
    if not files:
        print("  ❌ CONTRACT DETAIL file not found in the INPUT folder.")
        print(f"     Path: {INPUT_FOLDER}")
        print("     File name must start with: CONTRACT DETAIL")
        print("     Example: CONTRACT DETAIL 4.21.2026.xlsx")
        sys.exit(1)
    print(f"  📄 File: {files[0]}")
    return os.path.join(INPUT_FOLDER, files[0])


def detect_header_row(filepath, key_col):
    import pandas as pd
    df_raw = pd.read_excel(filepath, header=None, nrows=30, dtype=str)
    for i, row in df_raw.iterrows():
        vals = [str(v).strip() for v in row.values if str(v).strip() not in ("nan", "")]
        if key_col in vals:
            return i
    return 0


def read_contract_detail(filepath):
    import pandas as pd
    header_row = detect_header_row(filepath, COL_LEGACY_NUM)
    print(f"  📊 Header detected at row {header_row + 1}")
    df = pd.read_excel(filepath, header=header_row, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    required = [COL_LEGACY_NUM, COL_LEGACY_LINE, COL_TRANS_ID, COL_ITEM_NUM]
    missing = [c for c in required if c not in df.columns]
    if missing:
        print(f"\n  ❌ Missing columns: {missing}")
        print(f"     Available: {list(df.columns)}")
        print("     👉 Update column names in the CONFIGURATION section of STEP3.py")
        sys.exit(1)
    for col in required:
        df[col] = df[col].str.strip()
    df = df[
        df[COL_LEGACY_NUM].notna() &
        (df[COL_LEGACY_NUM] != "") &
        (df[COL_LEGACY_NUM].str.lower() != "nan")
    ].copy()
    print(f"  Valid rows in CONTRACT DETAIL: {len(df)}")
    return df


def expand_dashes(df):
    rows_out = []
    expanded = 0
    for _, row in df.iterrows():
        num_val  = str(row[COL_LEGACY_NUM]).strip()
        line_val = str(row[COL_LEGACY_LINE]).strip()
        if "-" in num_val:
            parts_num  = [p.strip() for p in num_val.split("-")]
            parts_line = [p.strip() for p in line_val.split("-")] if "-" in line_val else None
            for i, part in enumerate(parts_num):
                new_row = row.copy()
                new_row[COL_LEGACY_NUM] = part
                if parts_line and i < len(parts_line):
                    new_row[COL_LEGACY_LINE] = parts_line[i]
                rows_out.append(new_row)
            expanded += 1
        else:
            rows_out.append(row)
    import pandas as pd
    df_out = pd.DataFrame(rows_out).reset_index(drop=True)
    if expanded:
        print(f"  🔀 Rows expanded due to dashes: {expanded}")
        print(f"     Total rows after expansion: {len(df_out)}")
    return df_out


def to_num_str(v):
    try:
        return str(int(float(str(v).strip())))
    except (ValueError, TypeError):
        return str(v).strip()


def build_keys(df):
    df[COL_LEGACY_NUM]  = df[COL_LEGACY_NUM].apply(to_num_str)
    df[COL_LEGACY_LINE] = df[COL_LEGACY_LINE].apply(to_num_str)
    df["_KEY"] = df[COL_LEGACY_NUM] + "_" + df[COL_LEGACY_LINE]
    return df


def load_step2_data():
    if not os.path.exists(INTER_STEP2):
        print("\n  ⚠️  Step 2 data file not found (.step2_data.json)")
        print("      Make sure you ran STEP2.bat before this step.")
        sys.exit(1)
    with open(INTER_STEP2, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data.get("key_to_cfin", {}), data.get("date_label", "")


def merge_cfin(df, key_to_cfin):
    def get_cfin(key):
        return key_to_cfin.get(str(key).strip())
    df["CFIN DATE"] = df["_KEY"].apply(get_cfin)
    matched = df["CFIN DATE"].notna().sum()
    total   = len(df)
    print(f"  ✅ CFIN DATE merged: {matched}/{total} rows")
    if matched < total:
        unmatched = df[df["CFIN DATE"].isna()][COL_LEGACY_NUM].head(5).tolist()
        print(f"  ⚠️  No match for (first 5): {unmatched}")
    return df


# ─────────────────────────────────────────────────────────
# OWNER ASSIGNMENT — equitable round-robin by Transaction ID
# ─────────────────────────────────────────────────────────

def assign_owners(df):
    import pandas as pd

    output = df[[COL_TRANS_ID, COL_ITEM_NUM, "CFIN DATE"]].copy()
    output = output.rename(columns={
        COL_TRANS_ID: "TRANSACTION",
        COL_ITEM_NUM: "ITEM NUMBER IN DOCUMENT",
    })
    output = output.dropna(subset=["TRANSACTION"])

    unique_txs = list(dict.fromkeys(output["TRANSACTION"].tolist()))
    total_txs  = len(unique_txs)

    tx_owner = {}
    for i, tx in enumerate(unique_txs):
        tx_owner[tx] = OWNER_1 if i % 2 == 0 else OWNER_2

    output["OWNER"] = output["TRANSACTION"].map(tx_owner)

    h_count = sum(1 for v in tx_owner.values() if v == OWNER_1)
    d_count = sum(1 for v in tx_owner.values() if v == OWNER_2)
    print(f"  👥 Owner assignment: {OWNER_1}={h_count} TXs | {OWNER_2}={d_count} TXs  (total {total_txs} unique TXs)")

    return output



# ─────────────────────────────────────────────────────────
# RMR ACTIVATION.xlsx — guided download (no auth required)
# ─────────────────────────────────────────────────────────

def ensure_rmr_file(config):
    """
    Opens the SharePoint link in the browser and waits for the user
    to download the file. Detects it automatically once it appears.
    Removes any stale local copy beforehand.
    """
    import webbrowser
    import time

    rmr_url = config.get("rmr_activation_url", "")

    # Remove stale local copy so we always work with a fresh one
    if os.path.exists(LOCAL_RMR):
        os.remove(LOCAL_RMR)
        print("  🗑️  Removed previous local copy")

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    print("\n" + "=" * 60)
    print("  📥 DOWNLOAD RMR ACTIVATION.xlsx")
    print("=" * 60)
    print("\n  Opening SharePoint in your browser...")

    if rmr_url:
        webbrowser.open(rmr_url)
    else:
        print("  ⚠️  rmr_activation_url not set in config.json")
        print("      Open the file manually in your browser.")

    print()
    print("  In the browser:")
    print("  1. Click  File → Download  (or the Download button)")
    print(f"  2. Save it here with this exact name:")
    print(f"     {LOCAL_RMR}")
    print()
    print("  Waiting for the file to appear automatically...")
    print("  (press Ctrl+C to cancel and enter the path manually)")
    print("=" * 60)

    # Poll for up to 3 minutes, checking every 2 seconds
    timeout  = 180
    interval = 2
    elapsed  = 0

    try:
        while elapsed < timeout:
            if os.path.exists(LOCAL_RMR):
                # Small extra wait to ensure the file is fully written
                time.sleep(1)
                size_kb = os.path.getsize(LOCAL_RMR) // 1024
                print(f"\n  ✅ File detected! ({size_kb} KB) — continuing...")
                return
            time.sleep(interval)
            elapsed += interval
            if elapsed % 20 == 0:
                remaining = timeout - elapsed
                print(f"  ⏳ Still waiting... ({remaining}s remaining)")

    except KeyboardInterrupt:
        pass

    # Timeout or cancelled — ask for manual confirmation
    if not os.path.exists(LOCAL_RMR):
        print(f"\n  ⚠️  File not detected at expected path.")
        print(f"     Expected: {LOCAL_RMR}")
        input("\n  Save the file there and press ENTER to continue...")

        if not os.path.exists(LOCAL_RMR):
            print("\n  ❌ File still not found. Exiting.")
            input("\n  Press ENTER to close...")
            sys.exit(1)

# ─────────────────────────────────────────────────────────
# APPEND TO RMR ACTIVATION.xlsx
# ─────────────────────────────────────────────────────────

def get_current_month_sheet():
    from datetime import datetime
    return datetime.now().strftime("%B").upper()


def append_to_rmr_activation(new_rows_df):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Side
    from datetime import datetime

    sheet_name = get_current_month_sheet()
    wb = load_workbook(LOCAL_RMR)

    # ── Find sheet ────────────────────────────────────────
    if sheet_name not in wb.sheetnames:
        matches = [s for s in wb.sheetnames if sheet_name in s.upper()]
        if matches:
            sheet_name = matches[0]
            print(f"  ℹ️  Using sheet: '{sheet_name}'")
        else:
            print(f"  ❌ Sheet '{sheet_name}' not found.")
            print(f"      Available sheets: {wb.sheetnames}")
            sys.exit(1)
    else:
        print(f"  📋 Target sheet: '{sheet_name}'")

    ws = wb[sheet_name]

    # ── Locate header row ─────────────────────────────────
    EXPECTED_COLS = [
        "TRANSACTION ID",
        "ITEM NUMBER IN DOCUMENT",
        "CFIN DATE",
        "OWNER",
        "NOTES",
        "PROCESSING DATE",
        "COMPLETED",
    ]

    header_row_idx = None
    col_map = {}

    for row in ws.iter_rows(min_row=1, max_row=10):
        row_vals = [str(c.value).strip().upper() if c.value else "" for c in row]
        if "TRANSACTION ID" in row_vals:
            header_row_idx = row[0].row
            for cell in row:
                val = str(cell.value).strip().upper() if cell.value else ""
                for ec in EXPECTED_COLS:
                    if val == ec:
                        col_map[ec] = cell.column
            break

    if header_row_idx is None:
        print("  ❌ Could not find the header row in the sheet.")
        sys.exit(1)

    print(f"  📌 Header found at row {header_row_idx}")
    print(f"     Column mapping: {col_map}")

    all_col_indices = sorted(col_map.values())
    min_col = all_col_indices[0] if all_col_indices else 1
    max_col = all_col_indices[-1] if all_col_indices else 7

    # ── Find first empty row ──────────────────────────────
    tx_col = col_map.get("TRANSACTION ID", 1)
    last_data_row = header_row_idx
    for r in range(header_row_idx + 1, ws.max_row + 2):
        cell_val = ws.cell(row=r, column=tx_col).value
        if cell_val is None or str(cell_val).strip() == "":
            break
        last_data_row = r

    first_empty_row = last_data_row + 1
    print(f"  ➕ Appending {len(new_rows_df)} rows starting at row {first_empty_row}")

    # ── Border style — thin black, no fill, no center ─────
    border_side  = Side(style="thin", color="000000")
    cell_border  = Border(
        left=border_side, right=border_side,
        top=border_side,  bottom=border_side
    )
    left_align   = Alignment(horizontal="left",  vertical="center")
    right_align  = Alignment(horizontal="right", vertical="center")

    # ── Write rows ────────────────────────────────────────
    for i, (_, row_data) in enumerate(new_rows_df.iterrows()):
        target_row = first_empty_row + i

        def write_cell(col_key, value, align=left_align):
            if col_key not in col_map:
                return None
            c = ws.cell(row=target_row, column=col_map[col_key], value=value)
            c.border    = cell_border
            c.alignment = align
            return c

        # Transaction ID → number, right-aligned
        tx_val = row_data.get("TRANSACTION", "")
        try:
            tx_val = int(float(str(tx_val)))
        except (ValueError, TypeError):
            pass
        write_cell("TRANSACTION ID", tx_val, right_align)

        # Item number → number, right-aligned
        item_val = row_data.get("ITEM NUMBER IN DOCUMENT", "")
        try:
            item_val = int(float(str(item_val)))
        except (ValueError, TypeError):
            pass
        write_cell("ITEM NUMBER IN DOCUMENT", item_val, right_align)

        # CFIN DATE → date, left-aligned
        cfin_val = row_data.get("CFIN DATE", "")
        try:
            cfin_val = datetime.strptime(str(cfin_val), "%m/%d/%Y").date()
        except Exception:
            try:
                cfin_val = pd.to_datetime(cfin_val).date()
            except Exception:
                pass
        c = write_cell("CFIN DATE", cfin_val, left_align)
        if c:
            c.number_format = "MM/DD/YYYY"

        write_cell("OWNER",                  row_data.get("OWNER", ""), left_align)
        write_cell("NOTES",                  "",                         left_align)
        write_cell("PROCESSING DATE",        "",                         left_align)
        write_cell("COMPLETED",              "",                         left_align)

        # Apply border to any remaining columns in the table range
        for col_idx in range(min_col, max_col + 1):
            if col_idx not in col_map.values():
                cell = ws.cell(row=target_row, column=col_idx)
                cell.border    = cell_border
                cell.alignment = left_align

    # ── Extend existing table via XML ─────────────────────
    import zipfile, shutil, re

    wb.save(LOCAL_RMR)

    new_end_row = first_empty_row + len(new_rows_df) - 1
    tmp_path    = LOCAL_RMR + ".tmp.xlsx"
    shutil.copy(LOCAL_RMR, tmp_path)

    with zipfile.ZipFile(tmp_path, 'r') as zin, \
         zipfile.ZipFile(LOCAL_RMR, 'w', zipfile.ZIP_DEFLATED) as zout:

        table_patched = False
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith("xl/tables/") and item.filename.endswith(".xml"):
                text = data.decode("utf-8")

                def patch_ref(m):
                    return re.sub(
                        r'ref="([A-Z]+)(\d+):([A-Z]+)(\d+)"',
                        lambda m2: f'ref="{m2.group(1)}{m2.group(2)}:{m2.group(3)}{new_end_row}"',
                        m.group(0)
                    )

                new_text = re.sub(r'<table [^>]+>', patch_ref, text, count=1)
                data = new_text.encode("utf-8")
                table_patched = True
                print(f"  📐 Table XML patched → row {new_end_row}")
            zout.writestr(item, data)

    if not table_patched:
        print("  ⚠️  No table XML found — table not extended")

    os.remove(tmp_path)
    print(f"  💾 File saved locally.")
    return first_empty_row, new_end_row

# ─────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────

def main():
    banner("RMR ACTIVATION — STEP 3: CONTRACT DETAIL")

    print("\n⚙️  Loading configuration...")
    config = load_config()

    print("\n🔧 Searching for CONTRACT DETAIL file...")
    filepath = find_contract_file()

    print("\n🔄 Reading CONTRACT DETAIL...")
    df = read_contract_detail(filepath)

    print("\n🔀 Processing dashes in Legacy Contract Number...")
    df = expand_dashes(df)

    print("\n🔑 Building merge keys...")
    df = build_keys(df)

    print("\n🔗 Loading IW75 data (STEP 2)...")
    key_to_cfin, date_label = load_step2_data()

    print("\n📌 Merging with CFIN DATE...")
    df = merge_cfin(df, key_to_cfin)

    print("\n👥 Assigning owners (Hector / Daniel)...")
    new_rows = assign_owners(df)

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    # ── Auto-download RMR ACTIVATION.xlsx ────────────────
    ensure_rmr_file(config)

    print("\n📝 Appending new rows to RMR ACTIVATION.xlsx...")
    first_row, last_row = append_to_rmr_activation(new_rows)

    print("\n" + "=" * 60)
    print("  📋 MANUAL ACTION REQUIRED — UPLOAD TO SHAREPOINT:")
    print("=" * 60)
    print(f"\n  The updated file is at:")
    print(f"  {LOCAL_RMR}")
    print(f"\n  Upload it manually to SharePoint:")
    print(f"  1. Open Teams → the chat where RMR ACTIVATION.xlsx lives")
    print(f"  2. Click '...' on the file → Upload new version")
    print(f"     (or delete the old one and upload the new one)")
    print("=" * 60)

    banner(f"✅ PROCESS COMPLETED — {len(new_rows)} records appended")
    print(f"  Sheet:        {get_current_month_sheet()}")
    print(f"  Rows added:   {first_row} → {last_row}")
    h_rows = len(new_rows[new_rows["OWNER"] == OWNER_1])
    d_rows = len(new_rows[new_rows["OWNER"] == OWNER_2])
    print(f"  {OWNER_1}:   {h_rows} rows")
    print(f"  {OWNER_2}:  {d_rows} rows")
    print(f"  Output file:  {LOCAL_RMR}")
    print("=" * 60)

    print("\n  📂 Opening output folder...")
    try:
        os.startfile(OUTPUT_FOLDER)
    except Exception:
        pass

    print("\n  🎉 RMR Activation Step 3 completed!")
    input("\n  Press ENTER to close...")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nCancelled by user.")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        input("\nPress ENTER to close...")
        sys.exit(1)