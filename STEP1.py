#!/usr/bin/env python3
"""
============================================================
  RMR ACTIVATION — STEP 1: MONITORING
  Downloads the Monitoring file from SharePoint,
  applies filters and prepares SAP numbers for IW75.
============================================================
"""

import os
import sys
import json
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────
# CONFIGURATION — Update column names here if the
# Monitoring file structure changes
# ─────────────────────────────────────────────────────────

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE  = os.path.join(SCRIPT_DIR, "config.json")
INPUT_FOLDER = os.path.join(SCRIPT_DIR, "input")
INTER_FILE   = os.path.join(INPUT_FOLDER, ".step1_data.json")

# Exact column names in the Monitoring file
COL_CLOSE     = "Closed"
COL_INVOICE   = "Invoice"
COL_CFIN_DATE = "CFINDATE"
COL_SAP       = "SAP"

# ─────────────────────────────────────────────────────────

def banner(text):
    print("\n" + "=" * 60)
    print(f"  {text}")
    print("=" * 60)

def load_config():
    if not os.path.exists(CONFIG_FILE):
        print("⚠️  config.json not found. Using default values.")
        print("   Copy config.example.json to config.json and fill in your values.")
        return {}
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def install_if_missing(pkg):
    import importlib
    try:
        importlib.import_module(pkg)
    except ImportError:
        print(f"  📦 Installing {pkg}...")
        os.system(f"pip install {pkg} -q")

def check_dependencies():
    print("🔧 Checking dependencies...")
    for pkg in ["pandas", "openpyxl", "msal", "requests", "pyperclip"]:
        install_if_missing(pkg)
    print("  ✅ All dependencies installed\n")

def download_monitoring(config):
    return None  # Automatic download placeholder — configure credentials in config.json

    import msal
    import requests

    FILE_GUID  = config.get("file_guid", "")
    FILE_OWNER = config.get("file_owner_path", "")
    TENANT_ID  = config.get("tenant_id", "")
    SP_HOST    = config.get("sharepoint_host", "")
    CLIENT_ID  = "d3590ed6-52b3-4102-aeff-aad2292ab01c"
    CACHE_FILE = os.path.join(SCRIPT_DIR, ".token_cache.bin")
    SCOPES     = [f"https://{SP_HOST}/.default"]

    cache = msal.SerializableTokenCache()
    if os.path.exists(CACHE_FILE):
        try:
            cache.deserialize(open(CACHE_FILE, "r").read())
        except Exception:
            pass

    app = msal.PublicClientApplication(
        CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{TENANT_ID}",
        token_cache=cache
    )

    result   = None
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(SCOPES, account=accounts[0])

    if not result or "access_token" not in result:
        print("🔐 Signing in to Microsoft 365...")
        print("   A browser window will open. Sign in with your corporate account.\n")
        try:
            result = app.acquire_token_interactive(SCOPES)
        except Exception as e:
            print(f"  ⚠️  Interactive login error: {e}")
            return None

    if cache.has_state_changed:
        with open(CACHE_FILE, "w") as f:
            f.write(cache.serialize())

    if not result or "access_token" not in result:
        print(f"  ⚠️  Login failed: {result.get('error_description', 'unknown')}")
        return None

    token        = result["access_token"]
    download_url = (
        f"https://{SP_HOST}/personal/{FILE_OWNER}"
        f"/_api/web/GetFileById('{FILE_GUID}')/$value"
    )

    try:
        resp = requests.get(
            download_url,
            headers={
                "Authorization": f"Bearer {token}",
                "Accept": "application/octet-stream"
            },
            timeout=60
        )
        if resp.status_code == 200:
            from datetime import datetime
            month_name = datetime.now().strftime("%B")
            filename   = f"Monitoring {month_name}.xlsx"
            filepath   = os.path.join(INPUT_FOLDER, filename)
            with open(filepath, "wb") as f:
                f.write(resp.content)
            print(f"  ✅ Downloaded → {filename}")
            return filepath
        else:
            print(f"  ⚠️  HTTP error {resp.status_code} while downloading.")
            return None
    except Exception as e:
        print(f"  ⚠️  Network error: {e}")
        return None

def manual_download(config):
    url = config.get("sharepoint_url", "")
    print("\n" + "─" * 60)
    print("📌 MANUAL DOWNLOAD")
    print("─" * 60)
    print("Automatic access failed. Follow these steps:")
    print()
    print("  1. Open this link in your browser:")
    if url:
        print(f"     {url}")
    print()
    print("  2. Download the file:")
    print("     File → Download (or Ctrl+S)")
    print()
    print(f"  3. Save it in:")
    print(f"     {INPUT_FOLDER}")
    print()
    print("  4. File name:")
    from datetime import datetime
    month = datetime.now().strftime("%B")
    print(f"     Monitoring {month}.xlsx")
    print("─" * 60)
    input("\n  Press ENTER once you have saved the file...")

    files = sorted(
        [f for f in os.listdir(INPUT_FOLDER)
         if f.lower().startswith("monitoring") and f.endswith(".xlsx")],
        reverse=True
    )
    if files:
        fp = os.path.join(INPUT_FOLDER, files[0])
        print(f"  ✅ File found: {files[0]}")
        return fp
    else:
        print("  ❌ Monitoring file not found in INPUT folder.")
        sys.exit(1)

def get_sheets(filepath):
    from datetime import datetime, timedelta
    import openpyxl

    wb     = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheets = [s for s in wb.sheetnames if s.lower().startswith("monitoring")]
    wb.close()

    if not sheets:
        wb     = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()

    yesterday  = datetime.now() - timedelta(days=1)
    auto_sheet = f"Monitoring {yesterday.strftime('%m.%d.%Y')}"

    print()
    print("  Available sheets:")
    for i, s in enumerate(sheets):
        marker = "  ← yesterday" if s == auto_sheet else ""
        print(f"    [{i}] {s}{marker}")

    print()
    print("  You can select one or multiple days:")
    print("    Single day  →  0")
    print("    List        →  0,1,2")
    print("    Range       →  0-2")
    print("    All         →  all")
    print()

    while True:
        raw = input("  Which sheet(s) to process? ").strip().lower()

        if raw == "all":
            selected = list(range(len(sheets)))

        elif "-" in raw and "," not in raw:
            try:
                parts      = raw.split("-")
                start, end = int(parts[0].strip()), int(parts[1].strip())
                selected   = list(range(start, end + 1))
            except (ValueError, IndexError):
                print("  ⚠️  Invalid range. Example: 0-2")
                continue

        elif "," in raw:
            try:
                selected = [int(x.strip()) for x in raw.split(",")]
            except ValueError:
                print("  ⚠️  Invalid list. Example: 0,1,2")
                continue

        else:
            try:
                selected = [int(raw)]
            except ValueError:
                print("  ⚠️  Invalid input. Enter a number, list (0,1), range (0-2), or 'all'.")
                continue

        invalid = [i for i in selected if i < 0 or i >= len(sheets)]
        if invalid:
            print(f"  ⚠️  Index out of range: {invalid}. Valid range: 0–{len(sheets)-1}")
            continue

        chosen = [sheets[i] for i in selected]
        print()
        print(f"  ✅ Selected: {', '.join(chosen)}")
        return chosen

def process_monitoring(filepath):
    import pandas as pd

    print("\n🔄 Processing Monitoring...")
    sheets = get_sheets(filepath)

    all_frames = []
    for sheet in sheets:
        print(f"\n  📋 Sheet: {sheet}")
        df = pd.read_excel(filepath, sheet_name=sheet, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
        print(f"     Rows read: {len(df)}")

        required = [COL_CLOSE, COL_INVOICE, COL_CFIN_DATE, COL_SAP]
        missing  = [c for c in required if c not in df.columns]
        if missing:
            print(f"\n  ❌ Missing columns in sheet '{sheet}': {missing}")
            print(f"  Available columns: {list(df.columns)}")
            print("  👉 Update column names in the CONFIGURATION section of STEP1.py")
            sys.exit(1)

        df["_cfin_dt"] = pd.to_datetime(df[COL_CFIN_DATE], errors="coerce")
        today          = pd.Timestamp.today().normalize()

        cond_close   = df[COL_CLOSE].str.strip().str.upper() == "YES"
        cond_invoice = (
            df[COL_INVOICE].notna() &
            (df[COL_INVOICE].str.strip() != "") &
            (df[COL_INVOICE].str.strip().str.lower() != "nan")
        )
        cond_cfin = df["_cfin_dt"].isna() | (df["_cfin_dt"] <= today)

        filtered = df[(cond_close | cond_invoice) & cond_cfin].copy()
        print(f"     Rows after filtering: {len(filtered)}")

        result = filtered[[COL_SAP, "_cfin_dt"]].copy()
        result = result.rename(columns={"_cfin_dt": COL_CFIN_DATE})
        result[COL_SAP] = result[COL_SAP].str.strip()
        result = result[
            result[COL_SAP].notna() &
            (result[COL_SAP] != "") &
            (result[COL_SAP].str.lower() != "nan")
        ]
        result["_source_sheet"] = sheet
        all_frames.append(result)

    if not all_frames:
        print("\n  ❌ No data found in selected sheets.")
        sys.exit(1)

    combined = pd.concat(all_frames, ignore_index=True)

    before_dedup = len(combined)
    combined = combined.sort_values(COL_CFIN_DATE, na_position="last")
    combined = combined.drop_duplicates(subset=[COL_SAP], keep="first")
    after_dedup = len(combined)

    if before_dedup != after_dedup:
        print(f"\n  ℹ️  {before_dedup - after_dedup} duplicate SAP(s) across sheets — kept earliest CFIN DATE.")

    print(f"\n  ✅ Total unique SAP numbers: {len(combined)}")
    return combined

def parse_sheet_date(sheet_name):
    """
    Parses a sheet name like 'Monitoring 04.24.2026' into a datetime object.
    Returns None if parsing fails.
    """
    from datetime import datetime
    try:
        date_str = sheet_name.strip().split(" ", 1)[1]
        return datetime.strptime(date_str, "%m.%d.%Y")
    except Exception:
        return None

def build_date_label(sheet_names):
    """
    Builds a smart, human-readable date label from one or more sheet names.

    Rules:
      - 1 sheet          → "4.28.2026"
      - contiguous range → "4.26-28.2026"  (same month)
                         → "4.29-5.1.2026" (cross-month)
      - non-contiguous   → "4.24_4.26_4.28.2026"
      - unparseable      → "April" (month fallback)
    """
    from datetime import datetime, timedelta

    dates = []
    for s in sheet_names:
        dt = parse_sheet_date(s)
        if dt:
            dates.append(dt)

    if not dates:
        return datetime.now().strftime("%B")

    dates = sorted(set(dates))

    def fmt(dt):
        return f"{dt.month}.{dt.day}.{dt.year}"

    if len(dates) == 1:
        return fmt(dates[0])

    is_contiguous = all(
        (dates[i + 1] - dates[i]).days == 1
        for i in range(len(dates) - 1)
    )

    if is_contiguous:
        start, end = dates[0], dates[-1]
        if start.year == end.year and start.month == end.month:
            return f"{start.month}.{start.day}-{end.day}.{end.year}"
        else:
            return f"{start.month}.{start.day}-{end.month}.{end.day}.{end.year}"
    else:
        parts = [f"{dt.month}.{dt.day}" for dt in dates]
        return "_".join(parts) + f".{dates[-1].year}"

def extract_date_from_sheet(sheet_name):
    dt = parse_sheet_date(sheet_name)
    if dt:
        return f"{dt.month}.{dt.day}.{dt.year}"
    return None

def save_intermediate(result):
    import pandas as pd

    mapping = {}
    for _, row in result.iterrows():
        sap  = str(row[COL_SAP]).strip()
        cfin = row[COL_CFIN_DATE]
        mapping[sap] = cfin.strftime("%Y-%m-%d") if pd.notna(cfin) else None

    sheets_processed = (
        list(result["_source_sheet"].unique())
        if "_source_sheet" in result.columns else []
    )

    date_label = build_date_label(sheets_processed)

    first_sheet_date = None
    if sheets_processed:
        first_sheet_date = extract_date_from_sheet(sheets_processed[0])

    data = {
        "sap_to_cfin":      mapping,
        "sheets_processed": sheets_processed,
        "date_label":       date_label,
        "processing_date":  first_sheet_date
    }
    os.makedirs(os.path.dirname(INTER_FILE), exist_ok=True)
    with open(INTER_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

    print(f"\n  💾 Data saved for STEP2 ({len(mapping)} records)")
    print(f"  📅 Date label: {date_label}")

def copy_to_clipboard(text):
    try:
        import pyperclip
        pyperclip.copy(text)
        return True
    except Exception:
        return False

def main():
    banner("RMR ACTIVATION — STEP 1: MONITORING")
    check_dependencies()
    os.makedirs(INPUT_FOLDER, exist_ok=True)
    config = load_config()

    print("📥 Attempting automatic download from SharePoint...")
    filepath = download_monitoring(config)

    if not filepath:
        filepath = manual_download(config)

    result = process_monitoring(filepath)
    save_intermediate(result)

    sap_list  = result[COL_SAP].tolist()
    sap_block = "\n".join(sap_list)

    banner(f"SAP NUMBERS READY — {len(sap_list)} records")
    print(sap_block)
    print("=" * 60)

    if copy_to_clipboard(sap_block):
        print("\n📋 Numbers copied to clipboard ✅")
    else:
        print("\n⚠️  Could not copy automatically. Please copy manually.")

    sheets_processed = (
        list(result["_source_sheet"].unique())
        if "_source_sheet" in result.columns else []
    )
    date_label = build_date_label(sheets_processed)
    from datetime import datetime
    month = datetime.now().strftime("%B")

    banner("NEXT STEPS — SAP IW75")
    print("  1. Open SAP Fiori → transaction IW75")
    print("  2. 'Your Reference' field → multiple selection button")
    print("     (square icon on the right side of the field)")
    print("  3. In the multiple selection window:")
    print("     → Click the 'Clipboard' icon")
    print("     → Numbers will paste automatically")
    print("  4. Press F8 (Execute) to run the report")
    print("  5. Export to Excel:")
    print("     Menu → List → Save/Send → Local File → Spreadsheet")
    print(f"  6. Save in:")
    print(f"     {INPUT_FOLDER}")
    print(f"     Name: IW75 {month} {date_label}.xlsx")
    print()
    print("  7. Once you have the file → run STEP2.bat")
    print("=" * 60)
    input("\n  Press ENTER to close...")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nCancelled by user.")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        input("\nPress ENTER to close...")
        sys.exit(1)