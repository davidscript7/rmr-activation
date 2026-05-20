#!/usr/bin/env python3
"""
============================================================
RMR ACTIVATION — STEP 4: Excel → SharePoint Sync
Reads the Work sheet from RMR_Master.xlsx and syncs
changes to SharePoint Lista 2 and Lista 1.
============================================================
"""

import os
import sys
import json
import warnings
from datetime import datetime
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH  = os.path.join(SCRIPT_DIR, "config.json")

# Sheet names in RMR_Master.xlsx
WORK_SHEET = "Work"
LIST2_SHEET = "RMR ACTIVATION"

def banner(text):
    print("\n" + "=" * 60)
    print(f"  {text}")
    print("=" * 60)

def load_config():
    if not os.path.exists(CONFIG_PATH):
        print("  ❌ config.json not found.")
        sys.exit(1)
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def read_work_sheet(rmr_master_path):
    """Read items from Work sheet that need syncing"""
    import pandas as pd
    
    print(f"\n📂 Reading RMR_Master.xlsx...")
    print(f"   Path: {rmr_master_path}")
    
    if not os.path.exists(rmr_master_path):
        print(f"  ❌ File not found: {rmr_master_path}")
        sys.exit(1)
    
    try:
        df = pd.read_excel(rmr_master_path, sheet_name=WORK_SHEET, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
        print(f"  ✅ Sheet '{WORK_SHEET}' loaded: {len(df)} rows")
        
        # Filter only items pending sync
        if "Status" in df.columns:
            pending = df[df["Status"].str.strip() == "Pending sync"].copy()
            print(f"  🔄 Items pending sync: {len(pending)}")
            return pending
        else:
            print("  ⚠️  No 'Status' column found - syncing all items")
            return df
            
    except Exception as e:
        print(f"  ❌ Error reading file: {e}")
        sys.exit(1)

def connect_to_sharepoint(config):
    """Connect to SharePoint using Office365 REST API"""
    try:
        from office365.runtime.auth.client_credential import ClientCredential
        from office365.sharepoint.client_context import ClientContext
    except ImportError:
        print("\n  📦 Installing Office365-REST-Python-Client...")
        os.system("pip install Office365-REST-Python-Client --break-system-packages -q")
        from office365.runtime.auth.client_credential import ClientCredential
        from office365.sharepoint.client_context import ClientContext
    
    site_url = config.get("sharepoint_site_url")
    client_id = config.get("client_id", "")
    client_secret = config.get("client_secret", "")
    
    if not site_url:
        print("  ❌ sharepoint_site_url missing in config.json")
        sys.exit(1)
    
    print(f"\n🌐 Connecting to SharePoint...")
    print(f"   Site: {site_url}")
    
    try:
        if client_id and client_secret:
            credentials = ClientCredential(client_id, client_secret)
            ctx = ClientContext(site_url).with_credentials(credentials)
        else:
            # Use Windows integrated auth
            ctx = ClientContext(site_url)
        
        # Test connection
        web = ctx.web.get().execute_query()
        print(f"  ✅ Connected to: {web.title}")
        return ctx
        
    except Exception as e:
        print(f"  ❌ Connection failed: {e}")
        sys.exit(1)

def update_lista2_items(ctx, items_df, list_name):
    """Update items in SharePoint Lista 2"""
    
    print(f"\n📝 Updating SharePoint Lista 2: {list_name}")
    
    target_list = ctx.web.lists.get_by_title(list_name)
    
    updated_count = 0
    error_count = 0
    
    for idx, (_, row) in enumerate(items_df.iterrows(), 1):
        try:
            item_id = row.get("Id")
            if not item_id or str(item_id).strip() in ("", "nan", "0"):
                print(f"   ⚠️  Row {idx}: Missing SharePoint ID")
                error_count += 1
                continue
            
            item_id = int(float(str(item_id)))
            
            # Prepare update data
            update_data = {}
            
            # PROCESSED field
            processed = str(row.get("Processed", "")).strip().upper()
            if processed in ("YES", "NO"):
                update_data["PROCESSED"] = processed
            
            # PROCESSING DATE field
            proc_date = row.get("Processing Date")
            if proc_date and str(proc_date).strip() not in ("", "nan", "None"):
                try:
                    if isinstance(proc_date, str):
                        dt = datetime.strptime(proc_date, "%m/%d/%Y")
                    else:
                        dt = proc_date
                    update_data["PROCESSING_x0020_DATE"] = dt.isoformat()
                except Exception:
                    pass
            
            # PROJECT NUMBER field
            proj_num = row.get("Project Number")
            if proj_num and str(proj_num).strip() not in ("", "nan"):
                update_data["PROJECT_x0020_NUMBER"] = str(proj_num).strip()
            
            if update_data:
                item = target_list.get_item_by_id(item_id)
                item.set_property("__metadata", {"type": target_list.properties["ListItemEntityTypeFullName"]})
                
                for key, value in update_data.items():
                    item.set_property(key, value)
                
                item.update()
                ctx.execute_query()
                
                updated_count += 1
                
                if idx % 10 == 0:
                    print(f"   Progress: {idx}/{len(items_df)} items...")
            else:
                print(f"   ⚠️  Row {idx}: No data to update")
                error_count += 1
                
        except Exception as e:
            print(f"   ⚠️  Error updating item {idx} (ID: {item_id}): {e}")
            error_count += 1
    
    print(f"\n  ✅ Updated: {updated_count} items")
    if error_count > 0:
        print(f"  ⚠️  Errors: {error_count} items")
    
    return updated_count

def check_completed_projects(ctx, config):
    """
    Check Lista 2 for completed projects and update Lista 1.
    A project is complete when ALL its items have Processed=YES.
    """
    
    print(f"\n🔍 Checking for completed projects...")
    
    list2_name = config.get("list_2_name", "RMR ACTIVATION")
    list1_name = config.get("list_1_name", "Install Projects to Process (1070)")
    
    # Get all items from Lista 2
    list2 = ctx.web.lists.get_by_title(list2_name)
    items = list2.items.get().execute_query()
    
    print(f"   📊 Total items in Lista 2: {len(items)}")
    
    # Group by Project Number
    projects = {}
    for item in items:
        proj_num = item.properties.get("PROJECT_x0020_NUMBER", "")
        owner = item.properties.get("OWNER", "")
        processed = item.properties.get("PROCESSED", "")
        
        if not proj_num or str(proj_num).strip() in ("", "None"):
            continue
        
        proj_num = str(proj_num).strip()
        
        if proj_num not in projects:
            projects[proj_num] = {
                "items": [],
                "owners": []
            }
        
        projects[proj_num]["items"].append({
            "processed": str(processed).strip().upper() == "YES",
            "owner": str(owner).strip()
        })
        
        if owner:
            projects[proj_num]["owners"].append(str(owner).strip())
    
    print(f"   📦 Unique projects found: {len(projects)}")
    
    # Find completed projects
    completed_projects = []
    for proj_num, data in projects.items():
        if all(item["processed"] for item in data["items"]):
            # Determine majority owner
            owners = data["owners"]
            if owners:
                majority_owner = max(set(owners), key=owners.count)
            else:
                majority_owner = "Unknown"
            
            completed_projects.append({
                "project_number": proj_num,
                "owner": majority_owner,
                "item_count": len(data["items"])
            })
    
    print(f"   ✅ Completed projects: {len(completed_projects)}")
    
    if not completed_projects:
        print("      No projects to update in Lista 1")
        return 0
    
    # Update Lista 1
    print(f"\n📝 Updating SharePoint Lista 1: {list1_name}")
    
    list1 = ctx.web.lists.get_by_title(list1_name)
    updated_count = 0
    
    for proj_data in completed_projects:
        proj_num = proj_data["project_number"]
        owner = proj_data["owner"]
        
        try:
            # Find item in Lista 1 by Project Number
            query = f"<View><Query><Where><Eq><FieldRef Name='Project_x0020_Number'/><Value Type='Text'>{proj_num}</Value></Eq></Where></Query></View>"
            items = list1.get_items(query).execute_query()
            
            if len(items) > 0:
                item = items[0]
                
                # Update RMR Status and RMR Processed By
                item.set_property("RMR_x0020_Status", "RMR Processed")
                item.set_property("RMR_x0020_Processed_x0020_By", owner)
                item.update()
                ctx.execute_query()
                
                updated_count += 1
                print(f"   ✅ {proj_num} → Owner: {owner}")
            else:
                print(f"   ⚠️  {proj_num} not found in Lista 1")
                
        except Exception as e:
            print(f"   ⚠️  Error updating {proj_num}: {e}")
    
    print(f"\n  ✅ Lista 1 updated: {updated_count} projects marked as 'RMR Processed'")
    return updated_count

def mark_synced_in_excel(rmr_master_path, synced_count):
    """Mark items as Synced in the Work sheet"""
    import pandas as pd
    from openpyxl import load_workbook
    
    if synced_count == 0:
        return
    
    print(f"\n📝 Updating Work sheet in Excel...")
    
    try:
        wb = load_workbook(rmr_master_path)
        ws = wb[WORK_SHEET]
        
        # Find Status column
        status_col = None
        for col in range(1, 20):
            if str(ws.cell(1, col).value).strip().upper() == "STATUS":
                status_col = col
                break
        
        if not status_col:
            print("  ⚠️  Status column not found in Work sheet")
            return
        
        # Update rows with "Pending sync" → "Synced"
        updated = 0
        for row in range(2, ws.max_row + 1):
            cell_value = str(ws.cell(row, status_col).value).strip()
            if cell_value == "Pending sync":
                ws.cell(row, status_col).value = "Synced"
                ws.cell(row, status_col).fill = None  # Remove yellow fill
                from openpyxl.styles import PatternFill
                ws.cell(row, status_col).fill = PatternFill(start_color="C8FFC8", 
                                                             end_color="C8FFC8", 
                                                             fill_type="solid")  # Green
                updated += 1
        
        wb.save(rmr_master_path)
        print(f"  ✅ Marked {updated} items as 'Synced' in Excel")
        
    except Exception as e:
        print(f"  ⚠️  Could not update Excel: {e}")

def main():
    banner("RMR ACTIVATION — STEP 4: Sync Excel → SharePoint")
    
    print("\n⚙️  Loading configuration...")
    config = load_config()
    
    rmr_master_path = config.get("rmr_master_path")
    if not rmr_master_path:
        print("  ❌ rmr_master_path missing in config.json")
        sys.exit(1)
    
    # Read items pending sync from Work sheet
    items_df = read_work_sheet(rmr_master_path)
    
    if len(items_df) == 0:
        print("\n  ℹ️  No items pending sync.")
        banner("SYNC COMPLETE — Nothing to sync")
        input("\n  Press ENTER to close...")
        return
    
    # Connect to SharePoint
    ctx = connect_to_sharepoint(config)
    
    # Update Lista 2 (RMR ACTIVATION)
    list2_name = config.get("list_2_name", "RMR ACTIVATION")
    updated = update_lista2_items(ctx, items_df, list2_name)
    
    # Check and update completed projects in Lista 1
    completed = check_completed_projects(ctx, config)
    
    # Mark items as Synced in Excel
    mark_synced_in_excel(rmr_master_path, updated)
    
    banner("✅ SYNC COMPLETED")
    print(f"  Lista 2 updated: {updated} items")
    print(f"  Lista 1 updated: {completed} projects")
    print("=" * 60)
    
    print("\n  💡 Next: Refresh RMR_Master.xlsx (Ctrl+Alt+F5) to see updates")
    input("\n  Press ENTER to close...")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nCancelled by user.")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        input("\nPress ENTER to close...")
        sys.exit(1)